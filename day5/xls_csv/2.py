from openpyxl import Workbook
from openpyxl import load_workbook

objw = load_workbook('stat.xlsx')

objs = objw['mon']

datacell = objs['B3']
print("data =", datacell.value)

datacell = objs.cell(row=3, column=2)
print("data =", datacell.value)
