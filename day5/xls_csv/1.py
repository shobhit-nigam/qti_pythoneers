from openpyxl import Workbook
from openpyxl import load_workbook

objw = load_workbook('stat.xlsx')

print(objw.sheetnames)
