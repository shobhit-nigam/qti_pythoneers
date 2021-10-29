from openpyxl import Workbook
from openpyxl import load_workbook

objw = load_workbook('stat.xlsx')

objs = objw['mon']
 

for i in range(2, objs.max_row+1):
    print(objs.cell(row=i, column=1).value)
